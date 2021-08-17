import newspaper
import openpyxl
import os

a = [[1,2],[3,4]]
print(a)
a.clear()
print(a)

# cat_list：各个分类的网址，同时也是判断一篇文章类别的依据
cat_url_list = ["http://finance.people.com.cn","http://house.people.com.cn","http://edu.people.com.cn","http://scitech.people.com.cn","http://military.people.com.cn","http://auto.people.com.cn/","http://ent.people.com.cn","http://game.people.com.cn/","http://media.people.com.cn/","http://cpc.people.com.cn/","http://env.people.com.cn/","http://health.people.com.cn/"]
# ava_url_list：实际可以作为source的分类的网址
ava_url_list222 = ["http://finance.people.com.cn", "http://house.people.com.cn", "http://edu.people.com.cn", "http://military.people.com.cn", "http://cpc.people.com.cn/"]
ava_url_list = ["http://military.people.com.cn", "http://cpc.people.com.cn/"]
# channel_name：分类的名称
channel_name = ["财经","房产","教育","科技","军事","汽车","体育","游戏","娱乐","其他"]

# people = newspaper.build('http://www.people.com.cn/', language='zh', memoize_articles = False)
# print("新闻数量：", people.size())
# news_title = []
# news_text = []
# news = people.articles
#
# article_info = []
# for i in range (len(news)):         #以新闻链接长度为循环次数
#     paper = news[i]
#     try:
#         paper.download()
#         paper.parse()
#         print(paper.url,type(paper.url))
#         for cat_index in range(len(cat_list)):
#             if cat_list[cat_index] in paper.url:
#                 print("分类为：", cat_index)
#                 break
#         # print("标题：",paper.title)
#         # print("正文：",paper.text)
#         article_info.append(paper.text)
#         article_info.append(channelName[cat_index])
#         article_info.append(paper.title.split("--")[0].strip())
#         print(article_info)
#         article_info.clear()
#         # if "finance" in paper.url:
#         #     print("finance yes!")
#
#         # news_title.append(paper.title)
#         # news_text.append(paper.text)       # 将新闻正文以列表形式逐一储存
#     except:
#         # news_title.append('NULL')          # 如果无法访问，以NULL替代
#         # news_text.append('NULL')
#         continue

def check_news_invalid(news):
    return "index.html" in news.url or " " in news.url or "#liuyan" in news.url or len(news.text) < 20 or len(news.title) <= 5

def export2excel(news_info, excel_name):
    workbook = openpyxl.load_workbook(excel_name)
    # sheetnames = workbook.sheetnames
    cat = news_info[1]
    table = workbook[cat]
    table.append(news_info)
    workbook.save(excel_name)

def catch_news():
    # cnt1 = 0
    for cat_url in ava_url_list:
        # cnt1 += 1
        # if cnt1 >=5:
        #     break
        news_info = [] # 单篇新闻的内容、类别、标题，稍后会被追加到excel表格中
        is_valid = False # 该篇新闻是否在给定的分类里，初始值为False
        # 把当前的分类板块的url作为新闻源
        news_source = newspaper.build(cat_url, # 各个板块的链接
                                      language="zh", # 语言
                                      memoize_articles=False, # 是否要记忆已经访问过的文章
                                      fetch_images=False # 是否要获取图片
                                      )
        # 该分类板块下的所有新闻的网址
        news_list = news_source.articles
        # cnt2 = 0
        for news in news_list:
            try:
                # cnt2 += 1
                # if cnt2 >= 10:
                #     break
                news.download() # 下载新闻
                news.parse() # 解析新闻
                url = news.url
                # 去掉一些无效链接和重复的链接
                if "index.html" in url or " " in url or "#liuyan" in url or len(news.text) < 20 or len(news.title) <= 5:
                    continue
                for i in range(len(cat_url_list)):
                    if i >= 9:
                        cat_index = 9 # 把下标大于等于9的新闻归为“其他”
                    else:
                        cat_index = i
                    if cat_url_list[i] in url:
                        print("分类为：", channel_name[cat_index])
                        is_valid = True
                        break
                if is_valid: # 如果新闻的类别是我们所需要的
                    print(url)
                    news_info.append(news.text.replace("\n",""))
                    news_info.append(channel_name[cat_index])
                    news_info.append(news.title.split("--")[0].strip())
                    print(news_info)
                else:
                    print("该新闻不是所需要的：", url)
                # export2excel(news_info,"新闻文本分类算法样本集 - 副本.xlsx")
            except Exception as e:
                continue
            finally:
                news_info.clear()
                is_valid = False
        # print(news_source.size())
        # if news_source.size() != 0:
        #     print(cat_url, news_source.size())



if __name__ == "__main__":
    catch_news()