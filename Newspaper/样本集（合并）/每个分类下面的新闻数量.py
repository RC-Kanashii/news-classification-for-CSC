import openpyxl

workbook = openpyxl.load_workbook("新闻文本分类算法样本集+人民网+网易+旅兴网+读书网+美食台+美食天下+新浪+中华网+铁血+军事前沿+西陆+搜狗新闻库.xlsx")
sheetnames = workbook.sheetnames

cnt = 0
for cat in sheetnames:
    table = workbook[cat]
    row = table.max_row-1
    cnt += row
    print("{}下有{}条新闻".format(cat, row))
print("共有{}条新闻".format(cnt))