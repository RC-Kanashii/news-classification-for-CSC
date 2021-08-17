import openpyxl
import os
import jieba


# stopWord = [] # 停用词

def calFilename(num: int):
    maxlen = 6  # 文件名一共是6位
    t = num
    cnt = 0
    while t > 0:
        t //= 10
        cnt += 1
    filename = "0" * (maxlen - cnt) + str(num) + ".txt"
    return filename


def loadStopWord():
    with open("停用词表.txt", "r", encoding="utf-8") as file:
        stopword = file.read().splitlines()
        return stopword


def purifyWords(content: str, stopword):
    contentList = jieba.cut(content)
    res = ''
    for word in contentList:
        if word not in stopword:
            res += word
    return res

# HanLP.Config.ShowTermNature = False
# stopWord = loadStopWord()
# print(stopWord)
# content = "在一起涉企保全执行案中，最高法院力纠超标查封问题。5月19日，最高人民法院发布第三批《人民法院充分发挥审判职能作用保护产权和企业家合法权益典型案例》。"
# print(jieba.cut(content))
# print(purifyWords(content, stopWord))

# table = workbook.get_sheet_by_name("财经")
# print(table)
# rows = table.max_row
# cols = table.max_column
# print(rows)
# print(cols)
# print(type(table.cell(2, 1).value))
# table = workbook.get_sheet_by_name("财经")
# print(str(table.cell(133, 1).value))


if __name__ == "__main__":
    stopWord = loadStopWord()
    workbook = openpyxl.load_workbook(r"新闻文本分类算法样本集.xlsx")
    sheetnames = workbook.sheetnames
    try:
        os.makedirs(r".\新闻库（过滤停用词）\财经")
        os.makedirs(r".\新闻库（过滤停用词）\房产")
        os.makedirs(r".\新闻库（过滤停用词）\教育")
        os.makedirs(r".\新闻库（过滤停用词）\科技")
        os.makedirs(r".\新闻库（过滤停用词）\军事")
        os.makedirs(r".\新闻库（过滤停用词）\汽车")
        os.makedirs(r".\新闻库（过滤停用词）\体育")
        os.makedirs(r".\新闻库（过滤停用词）\游戏")
        os.makedirs(r".\新闻库（过滤停用词）\娱乐")
        os.makedirs(r".\新闻库（过滤停用词）\其他")
    except Exception as E:
        print(E)
    for cat in sheetnames:
        table = workbook[cat]
        max_row = table.max_row
        max_col = table.max_column
        for row in range(2, max_row + 1):
            text1 = str(table.cell(row, 3).value)
            text2 = str(table.cell(row, 1).value)
            if text1 == "None":
                text1 = ""
            if text2 == "None":
                text2 = ""
            content = text1 + "\n" + text2
            content = purifyWords(content,stopWord)
            filename = r".\新闻库（过滤停用词）" + '\\' + cat + '\\' + calFilename(row - 1)
            with open(filename, "w", encoding="utf-8") as f:
                f.write(content)