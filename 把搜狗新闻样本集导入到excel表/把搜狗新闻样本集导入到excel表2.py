import openpyxl
import re

def calFilename(num: int):
    maxlen = 4  # 文件名一共是4位
    t = num
    cnt = 0
    while t > 0:
        t //= 10
        cnt += 1
    filename = "0" * (maxlen - cnt) + str(num) + ".txt"
    return filename

if __name__ == "__main__":
    workbook = openpyxl.load_workbook(r"新闻文本分类算法样本集 - 空2.xlsx")
    # print(calFilename(10030))
    sheetnames = ["教育", "军事", "其他", "汽车", "体育"]
    for cat in sheetnames:
        table = workbook[cat]
        print("正在保存", cat, "下的新闻……")
        for i in range(1,1001):
            filename = r".\搜狗文本分类语料库迷你版" + "\\" + cat + "\\" + calFilename(i)
            # print(filename)
            try:
                with open(filename, "r", encoding="utf-8") as f:
                    text = f.read().replace("\n", "").replace("　","")
                    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
                    text = ILLEGAL_CHARACTERS_RE.sub(r'', text)
                    # print(text)
                    print("{}下的第{}篇新闻".format(cat, i))
                    info = ["", cat, text]
                    table.append(info)
                    if i % 100 == 0:
                        workbook.save("新闻文本分类算法样本集 - 空2.xlsx")
                        print("保存中")
            except:
                continue
            # for row in range(2, max_row + 1):
        #     text1 = str(table.cell(row, 3).value)
        #     text2 = str(table.cell(row, 1).value)
        #     if text1 == "None":
        #         text1 = ""
        #     if text2 == "None":
        #         text2 = ""
        #     content = text1 + "\n" + text2
        #     # content = purifyWords(content,stopWord) # 过滤停用词
        #     filename = r".\新闻库 新闻文本分类算法样本集+人民网+网易+旅兴网+读书网+美食台+美食天下" + '\\' + cat + '\\' + calFilename(row - 1)
        #     with open(filename, "w", encoding="utf-8") as f:
        #         f.write(content)
        print(cat, "下的新闻保存完毕！")