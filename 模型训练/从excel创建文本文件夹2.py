# coding:utf-8

import re
import openpyxl
import os
import jieba


# stopWord = [] # 停用词

def calFilename(num: int):
    maxlen = 6  # 文件名一共是6位
    t = num
    cnt = 0
    while t > 0:
        t //= 10
        cnt += 1
    filename = "0" * (maxlen - cnt) + str(num) + ".txt"
    return filename


def loadStopWord():
    with open("停用词表.txt", "r", encoding="utf-8") as file:
        stopword = file.read().splitlines()
        return stopword


def purifyWords(content: str, stopword):
    contentList = jieba.cut(content)
    res = ''
    for word in contentList:
        if word not in stopword:
            res += word
    return res

def filter(text):
    # 正则过滤掉特殊符号、标点、英文、数字等。
    pattern = '[a-zA-Z0-9’!"#$%&\'()*+,-./:：;；|<=>?@，—。?★、…【】《》？“”‘’！[\\]^_`{|}~（）]+'
    # 去除换行符
    text=re.sub(pattern, ' ', text)
    # 多个空格成1个
    text=text.replace(" ", "")
    return text


if __name__ == "__main__":
    # stopWord = loadStopWord()
    excel_name = r"新闻文本分类算法样本集+人民网+网易+旅兴网+读书网+美食台+美食天下+新浪+中华网+铁血+军事前沿+西陆+搜狗新闻库+数码科技网+科技快报网"
    prefix = ""#"过滤标点、英语、数字 "
    workbook = openpyxl.load_workbook(excel_name + ".xlsx")
    sheetnames = workbook.sheetnames
    try:
        os.makedirs(r"."+ "\\" + prefix + excel_name + "\财经")
        os.makedirs(r"."+ "\\" + prefix + excel_name + "\房产")
        os.makedirs(r"."+ "\\" + prefix + excel_name + "\教育")
        os.makedirs(r"."+ "\\" + prefix + excel_name + "\科技")
        os.makedirs(r"."+ "\\" + prefix + excel_name + "\军事")
        os.makedirs(r"."+ "\\" + prefix + excel_name + "\汽车")
        os.makedirs(r"."+ "\\" + prefix + excel_name + "\体育")
        os.makedirs(r"."+ "\\" + prefix + excel_name + "\游戏")
        os.makedirs(r"."+ "\\" + prefix + excel_name + "\娱乐")
        os.makedirs(r"."+ "\\" + prefix + excel_name + "\其他")
    except Exception as E:
        print(E)
    for cat in sheetnames:
        table = workbook[cat]
        max_row = table.max_row
        max_col = table.max_column
        print("正在创建", cat, "下的新闻……")
        for row in range(2, max_row + 1):
            text1 = str(table.cell(row, 3).value)
            text2 = str(table.cell(row, 1).value)
            if text1 == "None":
                text1 = ""
            if text2 == "None":
                text2 = ""
            content = text1 + "\n" + text2
            # content = purifyWords(content,stopWord) # 过滤停用词
            # content = filter(content) # 过滤标点符号、英文、数字
            filename = r"."+ "\\" + prefix + excel_name + '\\' + cat + '\\' + calFilename(row - 1)
            with open(filename, "w", encoding="utf-8") as f:
                f.write(content)
        print(cat, "下的新闻创建完毕！")