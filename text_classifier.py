import os
import re
from pyhanlp import SafeJClass
import openpyxl

model_path = "svm_2gram.ser"

LinearSVMClassifier = SafeJClass('com.hankcs.hanlp.classification.classifiers.LinearSVMClassifier')
IOUtil = SafeJClass('com.hankcs.hanlp.corpus.io.IOUtil')


def load_model(model_path):
    if os.path.isfile(model_path):
        return LinearSVMClassifier(IOUtil.readObjectFrom(model_path))


def classify(title, content, classifier):
    text = title + "\n" + content
    res = "该新闻属于：【{}】".format(classifier.classify(text)) + "\n属于各分类的概率为：\n"
    predict_dict = classifier.predict(text)
    for cat in predict_dict:
        res = res + cat + "：" + str(round(predict_dict[cat] * 100, 2)) + "%\n"
    return res


def act1():
    print("act123456")


def classify_all(file_dir, classifier):
    """
    对测试集中的新闻进行分类
    :param file_dir: 测试集的目录
    :param classifier: 分类器
    :param act1: 额外的操作（一般用于输出进度）
    :param act2: 额外的操作（一般用于输出“已完成”信息）
    :return: True
    """
    workbook = openpyxl.load_workbook(file_dir)
    if "类别" in workbook.sheetnames:
        table = workbook["类别"]
    else:
        table = workbook.active
    max_row = table.max_row
    for row in range(2, max_row + 1):
        title = table.cell(row, 3).value
        content = table.cell(row, 4).value
        if title == "None":
            title = ""
        if content == "None":
            content = ""
        text = title + "\n" + content
        # 去除空白字符
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]|\s+|\n')
        text = ILLEGAL_CHARACTERS_RE.sub(r'', text)
        # print(text)
        # 进行分类
        cat = classifier.classify(text)
        table.cell(row, 2).value = cat
    workbook.save(file_dir)
    print("全部保存完毕")
    return True


if __name__ == "__main__":
    classifier = load_model(model_path)
    print(classify("科技手机手机", "我国卫星手机电脑", classifier))
    classify_all("少量测试集.xlsx", classifier)